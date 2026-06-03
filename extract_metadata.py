import os
import subprocess
import pandas as pd
import openpyxl

def sync_keep_format(target_directory, output_file):
    cmd = ['exiftool', '-r', '-csv', '-Artist', '-Album', '-Title', '-ext', 'mp3', target_directory]
    result_bytes = subprocess.check_output(cmd, stderr=subprocess.STDOUT)
    result_str = result_bytes.decode('utf-8', errors='replace')
    
    new_df = pd.read_csv(pd.io.common.StringIO(result_str))
    if 'SourceFile' in new_df.columns: new_df = new_df.drop(columns=['SourceFile'])
    new_df = new_df.reindex(columns=['Artist', 'Album', 'Title'])

    # Modify existing file
    if os.path.exists(output_file):
        book = openpyxl.load_workbook(output_file)
        sheet = book.active
        
        # Only delete rows 2 onwards to protect header row (Row 1)
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row)
        
        for row in new_df.values.tolist():
            sheet.append(row)
            
        book.save(output_file)
    else:
        new_df.to_excel(output_file, index=False)

    print(f"Sync complete. Formatting and headers preserved in: {output_file}")

if __name__ == '__main__':
    sync_keep_format('/home/alelawso/Music', '/home/alelawso/Desktop/Music_Library.xlsx')